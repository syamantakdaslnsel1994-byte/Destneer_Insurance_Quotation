#!/usr/bin/env python3
"""
Full structural analysis of an Excel file using openpyxl.
Extracts: sheet names, cell values/formulas, fonts, fills, alignment,
borders, merged cells, column widths, row heights.
"""

import sys
import json
try:
    import openpyxl
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Installing openpyxl...")
    import subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.utils import get_column_letter

def safe_color(color_obj):
    """Extract ARGB string from a color object safely."""
    if color_obj is None:
        return None
    try:
        if color_obj.type == "rgb":
            return color_obj.rgb
        elif color_obj.type == "indexed":
            return f"indexed:{color_obj.indexed}"
        elif color_obj.type == "theme":
            return f"theme:{color_obj.theme},tint:{color_obj.tint}"
        else:
            return str(color_obj)
    except Exception:
        return str(color_obj)

def safe_border_side(side):
    """Extract border side info."""
    if side is None:
        return None
    try:
        return {
            "style": side.style,
            "color": safe_color(side.color) if side.color else None
        }
    except Exception:
        return None

def analyze_file(filepath):
    report = {}

    # Load workbook (formula mode)
    wb = openpyxl.load_workbook(filepath, data_only=False)

    # 1. Sheet names
    report["sheet_names"] = wb.sheetnames
    report["sheets"] = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        sheet_data = {}

        # 2. Merged cell ranges
        merged_ranges = [str(mr) for mr in ws.merged_cells.ranges]
        sheet_data["merged_cell_ranges"] = merged_ranges

        # Build a lookup: which merged range does each cell belong to?
        merged_lookup = {}
        for mr in ws.merged_cells.ranges:
            range_str = str(mr)
            for row in ws.iter_rows(min_row=mr.min_row, max_row=mr.max_row,
                                     min_col=mr.min_col, max_col=mr.max_col):
                for cell in row:
                    merged_lookup[cell.coordinate] = range_str

        # 3. Column widths
        col_widths = {}
        for col_letter, col_dim in ws.column_dimensions.items():
            col_widths[col_letter] = {
                "width": col_dim.width,
                "custom_width": col_dim.customWidth,
                "hidden": col_dim.hidden,
                "auto_size": col_dim.auto_size
            }
        sheet_data["column_widths"] = col_widths

        # 4. Row heights
        row_heights = {}
        for row_num, row_dim in ws.row_dimensions.items():
            row_heights[row_num] = {
                "height": row_dim.height,
                "custom_height": row_dim.customHeight,
                "hidden": row_dim.hidden
            }
        sheet_data["row_heights"] = row_heights

        # 5. Cell data
        cells = {}
        for row in ws.iter_rows():
            for cell in row:
                # Skip truly empty cells (no value, no formula, no special formatting)
                # But we want to capture cells with formatting even if value is None
                has_value = cell.value is not None
                has_font = cell.font is not None
                has_fill = cell.fill is not None and cell.fill.fill_type not in (None, "none")
                has_alignment = cell.alignment is not None
                has_border = cell.border is not None

                # Check if any border side has a style
                border_has_style = False
                if cell.border:
                    for side in [cell.border.top, cell.border.bottom, cell.border.left, cell.border.right]:
                        if side and side.style:
                            border_has_style = True
                            break

                # Check if font is non-default
                font_non_default = False
                if cell.font:
                    f = cell.font
                    if f.bold or f.italic or f.underline or (f.size and f.size != 11) or f.color:
                        if f.color and safe_color(f.color) not in (None, "FF000000", "00000000"):
                            font_non_default = True
                        elif f.bold or f.italic or f.underline or (f.size and f.size != 11):
                            font_non_default = True

                # Check alignment
                align_non_default = False
                if cell.alignment:
                    a = cell.alignment
                    if a.horizontal or a.vertical or a.wrap_text:
                        align_non_default = True

                # Include cell if it has value OR notable formatting
                if not (has_value or has_fill or border_has_style or font_non_default or align_non_default or cell.coordinate in merged_lookup):
                    continue

                cell_info = {
                    "row": cell.row,
                    "column": cell.column,
                    "address": cell.coordinate,
                    "value": str(cell.value) if cell.value is not None else None,
                    "data_type": cell.data_type,
                    "number_format": cell.number_format,
                    "in_merged_range": merged_lookup.get(cell.coordinate),
                }

                # Font
                if cell.font:
                    f = cell.font
                    cell_info["font"] = {
                        "name": f.name,
                        "size": f.size,
                        "bold": f.bold,
                        "italic": f.italic,
                        "underline": f.underline,
                        "strike": f.strike,
                        "color_argb": safe_color(f.color),
                        "vertAlign": f.vertAlign,
                    }
                else:
                    cell_info["font"] = None

                # Fill
                if cell.fill:
                    fill = cell.fill
                    cell_info["fill"] = {
                        "fill_type": fill.fill_type,
                        "fg_color_argb": safe_color(fill.fgColor),
                        "bg_color_argb": safe_color(fill.bgColor),
                    }
                else:
                    cell_info["fill"] = None

                # Alignment
                if cell.alignment:
                    a = cell.alignment
                    cell_info["alignment"] = {
                        "horizontal": a.horizontal,
                        "vertical": a.vertical,
                        "wrap_text": a.wrap_text,
                        "shrink_to_fit": a.shrinkToFit,
                        "indent": a.indent,
                        "text_rotation": a.textRotation,
                    }
                else:
                    cell_info["alignment"] = None

                # Border
                if cell.border:
                    b = cell.border
                    cell_info["border"] = {
                        "top": safe_border_side(b.top),
                        "bottom": safe_border_side(b.bottom),
                        "left": safe_border_side(b.left),
                        "right": safe_border_side(b.right),
                        "diagonal": safe_border_side(b.diagonal),
                        "diagonal_direction": getattr(b, "diagonal_direction", None),   # renamed in openpyxl 3.x
                    }
                else:
                    cell_info["border"] = None

                cells[cell.coordinate] = cell_info

        sheet_data["cells"] = cells
        report["sheets"][sheet_name] = sheet_data

    wb.close()
    return report

def print_report(report):
    print("=" * 80)
    print("EXCEL FILE STRUCTURE REPORT")
    print("=" * 80)

    print(f"\nSHEET NAMES: {report['sheet_names']}")
    print(f"TOTAL SHEETS: {len(report['sheet_names'])}\n")

    for sheet_name, sheet_data in report["sheets"].items():
        print(f"\n{'='*80}")
        print(f"SHEET: '{sheet_name}'")
        print(f"{'='*80}")

        # Merged ranges
        print(f"\n--- MERGED CELL RANGES ({len(sheet_data['merged_cell_ranges'])}) ---")
        for mr in sheet_data["merged_cell_ranges"]:
            print(f"  {mr}")

        # Column widths
        print(f"\n--- COLUMN WIDTHS ---")
        for col, info in sorted(sheet_data["column_widths"].items()):
            print(f"  Col {col}: width={info['width']}, custom={info['custom_width']}, hidden={info['hidden']}")

        # Row heights
        print(f"\n--- ROW HEIGHTS (rows with custom/explicit heights) ---")
        for row_num in sorted(sheet_data["row_heights"].keys()):
            info = sheet_data["row_heights"][row_num]
            print(f"  Row {row_num}: height={info['height']}, custom={info['custom_height']}, hidden={info['hidden']}")

        # Cells
        print(f"\n--- CELLS WITH CONTENT/FORMATTING ({len(sheet_data['cells'])}) ---")
        # Sort cells by row then column
        def cell_sort_key(addr):
            import re
            m = re.match(r"([A-Z]+)(\d+)", addr)
            if m:
                col_str, row_str = m.group(1), m.group(2)
                # Convert column letters to number
                col_num = 0
                for ch in col_str:
                    col_num = col_num * 26 + (ord(ch) - ord('A') + 1)
                return (int(row_str), col_num)
            return (0, 0)

        for addr in sorted(sheet_data["cells"].keys(), key=cell_sort_key):
            c = sheet_data["cells"][addr]
            print(f"\n  CELL {addr} (Row {c['row']}, Col {c['column']})")
            print(f"    Value       : {c['value']}")
            print(f"    Data Type   : {c['data_type']}")
            print(f"    Num Format  : {c['number_format']}")
            print(f"    Merged Range: {c['in_merged_range']}")

            if c["font"]:
                f = c["font"]
                print(f"    Font        : name={f['name']}, size={f['size']}, bold={f['bold']}, "
                      f"italic={f['italic']}, underline={f['underline']}, "
                      f"color={f['color_argb']}, strike={f['strike']}")

            if c["fill"]:
                fill = c["fill"]
                if fill["fill_type"] and fill["fill_type"] != "none":
                    print(f"    Fill        : type={fill['fill_type']}, fg={fill['fg_color_argb']}, bg={fill['bg_color_argb']}")

            if c["alignment"]:
                a = c["alignment"]
                if a["horizontal"] or a["vertical"] or a["wrap_text"]:
                    print(f"    Alignment   : h={a['horizontal']}, v={a['vertical']}, "
                          f"wrap={a['wrap_text']}, indent={a['indent']}, rotation={a['text_rotation']}")

            if c["border"]:
                b = c["border"]
                sides = []
                for side_name in ["top", "bottom", "left", "right"]:
                    side = b[side_name]
                    if side and side.get("style"):
                        sides.append(f"{side_name}:{side['style']}({side.get('color','')})")
                if sides:
                    print(f"    Border      : {', '.join(sides)}")

    print("\n" + "=" * 80)
    print("END OF REPORT")
    print("=" * 80)

if __name__ == "__main__":
    import sys
    filepath = sys.argv[1] if len(sys.argv) > 1 else r"Vivek Bhaia_Quote.xlsx"

    print(f"Analyzing: {filepath}")
    report = analyze_file(filepath)

    # Save JSON report
    json_out  = sys.argv[2] if len(sys.argv) > 2 else "excel_report.json"
    with open(json_out, "w", encoding="utf-8") as f:
        json.dump(report, f, indent=2, default=str, ensure_ascii=False)
    print(f"\nJSON report saved to: {json_out}")

    print_report(report)
